import openpyxl as xl
from datetime import datetime
from scraper import Scraper
from typing import Dict, Union

class FileWriter:
    def __init__(self):
        self.__index_to_month = {
            1: "Січень",
            2: "Лютий",
            3: "Березень",
            4: "Квітень",
            5: "Травень",
            6: "Червень",
            7: "Липень",
            8: "Серпень",
            9: "Вересень",
            10: "Жовтень",
            11: "Листопад",
            12: "Грудень"
        }

        self.__current_year = datetime.now().year
        self.__current_month = self.__index_to_month[datetime.now().month]

        self.__goal_workbook = xl.Workbook()
        self.__goal_worksheet = self.__goal_workbook.active
        self.__goal_worksheet.title = f"{self.__current_month} {self.__current_year}"
        self.__goal_worksheet.append(["Ім'я", "Виконані цілі", "Заохочення",  "Не виконані цілі", "Всього цілей"])

        self.__meas_workbook = xl.Workbook()
        self.__meas_worksheet = self.__meas_workbook.active
        self.__meas_worksheet.title = f"{self.__current_month} {self.__current_year}"
        self.__meas_worksheet.append(["Ім'я", "Вага", "Плечі",  "Груди", "Рука права", "Рука ліва", "Талія", "Стегна", "Стегна праве", "Стегна ліве"])
        

    def write_goal_data(self, data):
        self.__goal_worksheet.append(data)
        self.__goal_workbook.save("Report.xlsx")


    def write_measurement_data(self, data: Dict[str, Union[str, float]]):
        last_row = len(self.__meas_worksheet["A"])
        for k, v in data.items():
            for col in range(1, 11):
                if self.__meas_worksheet.cell(row=1, column=col).value == k:
                    self.__meas_worksheet.cell(row=last_row+1, column=col).value = v
                    
        self.__meas_workbook.save("Measurements.xlsx")




        

