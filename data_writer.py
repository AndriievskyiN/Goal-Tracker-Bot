import psycopg2
import openpyxl as xl
from openpyxl.styles import Font

from typing import Dict, Union, Tuple
from datetime import datetime

from hidden import hostname, database, username, pwd, port_id

class DataWriter:
    def __init__(self):
        self.__index_to_month = {
            1: "Січень",
            2: "Лютий",
            3: "Березень",
            4: "Квітень",
            5: "Травень",
            6: "Червень",
            7: "Липень",
            8: "Серпень",
            9: "Вересень",
            10: "Жовтень",
            11: "Листопад",
            12: "Грудень"
        }

        self.__conn = psycopg2.connect(
            host=hostname,
            dbname=database,
            user=username,
            password=pwd,
            port=port_id
        )
        self.__cur = self.__conn.cursor()

        # self.__cur.execute(
        #     """
        #     DROP TABLE goals
        #     """
        # )

        # self.__cur.execute(
        #     """
        #     DROP TABLE measurements
        #     """
        # )

        # self.__conn.commit()

        # create_goals_table = """
        #     CREATE TABLE IF NOT EXISTS goals (
        #         id SERIAL PRIMARY KEY,
        #         year SMALLINT,
        #         month SMALLINT,
        #         week_num SMALLINT,
        #         name VARCHAR(100) NOT NULL,
        #         completed_goals SMALLINT,
        #         rewards SMALLINT, 
        #         uncompleted_goals SMALLINT,
        #         total_goals SMALLINT,
        #         UNIQUE (year, month, week_num, name)
        #     )    
        # """

        # create_measurements_table = """
        #     CREATE TABLE IF NOT EXISTS measurements (
        #         id SERIAL PRIMARY KEY,
        #         year SMALLINT,
        #         month SMALLINT,
        #         week_num SMALLINT,
        #         name VARCHAR(100) NOT NULL,
        #         weight REAL,
        #         shoulders REAL,
        #         chest REAL,
        #         right_arm REAL, 
        #         left_arm REAL,
        #         waist REAL,
        #         hips REAL, 
        #         right_hip REAL,
        #         left_hip REAL
        #     )
        # """

        # self.__cur.execute(create_goals_table)
        # self.__cur.execute(create_measurements_table)
        # self.__conn.commit()

        # self.__current_year = datetime.now().year
        # self.__current_month = self.__index_to_month[datetime.now().month]

        # self.__meas_workbook = xl.Workbook()
        # self.__meas_worksheet = self.__meas_workbook.active
        # self.__meas_worksheet.title = f"{self.__current_month} {self.__current_year}"
        # self.__meas_worksheet.append(["Ім'я", "Вага", "Плечі",  "Груди", "Рука права", "Рука ліва", "Талія", "Стегна", "Стегна праве", "Стегна ліве"])
        
    def write_goals_xl(self, data):
        self.__goal_worksheet.append(data)
        self.__goal_workbook.save("Report.xlsx")


    def write_measurements_xl(self, data: Dict[str, Union[str, float]]):
        last_row = len(self.__meas_worksheet["A"])
        for k, v in data.items():
            for col in range(1, 11):
                if self.__meas_worksheet.cell(row=1, column=col).value == k:
                    self.__meas_worksheet.cell(row=last_row+1, column=col).value = v
                    
        self.__meas_workbook.save("Measurements.xlsx")


    
    def write_goal_data_db(self, data: Tuple[Union[str, int]]) -> str:
        self.__cur.execute(
        """
        INSERT INTO goals (year, month, week_num, name, completed_goals, rewards, uncompleted_goals, total_goals)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        
        ON CONFLICT (year, month, week_num, name)
            DO UPDATE
                SET 
                    completed_goals = excluded.completed_goals, 
                    rewards = excluded.rewards,
                    uncompleted_goals = excluded.uncompleted_goals,
                    total_goals = excluded.total_goals

        RETURNING 1
        """, 
        data)
        self.__conn.commit()
    
    # def TEST_write_measurement_data(self, data):
    #     # WILL NOT WORK (DATA IS SUPPOSED TO BE A DICTIONARY, BUT IN THIS FUNCTION IT IS TREATED AS A LIST)
    #     query = """
    #         INSERT INTO measurements (date, name, weight, shoulders, chest, right_arm, left_arm, waist, hips, right_hip, left_hip)
    #             VALUES (
    #                 %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
    #             )
    #     """

    #     insert_values = tuple(data)
        
    #     self.__cur.execute(query, insert_values)
    #     self.__conn.commit()


    def get_goals_xl(self, mode: str, sort_by: str):
        sort_by_options = ["completed_goals", "rewards", "uncompleted_goals", "total_goals"]
        sort_by_direction = "ASC" if sort_by == "uncompleted_goals" else "DESC"
        today = datetime.today().date()
        year = int(datetime.now().strftime("%Y"))
        month = int(datetime.now().strftime("%m"))
        month_str = self.__index_to_month[month]
        week_num = week_number_of_month(today)

        if mode == "week":
            filename = f"{year} {month_str} {week_num}.xlsx"

            query = """
                SELECT 
                    name, completed_goals, rewards, uncompleted_goals, total_goals
                FROM
                    goals
                WHERE year = %s and month = %s and week_num = %s
            """

            final_weekly_query = """
                SELECT 
                    sum(completed_goals) as completed_goals , sum(rewards) as rewards, sum(uncompleted_goals) as uncompleted_goals, sum(total_goals) as total_goals
                FROM 
                    goals
                WHERE 
                    year = %s and month = %s and week_num = %s
                GROUP BY
                    week_num, month, year
            """

            insert_values = (year, month, week_num)

            self.__cur.execute(query, insert_values)
            data = self.__cur.fetchall()

            self.__cur.execute(final_weekly_query, insert_values)
            final_weekly_data = self.__cur.fetchall()

            if data:
                # Get index to sort by
                sort_by_index = sort_by_options.index(sort_by) + 1 # + 1 because we exclude name 

                if sort_by == "uncompleted_goals":
                    data = sorted(data, key=lambda x: x[sort_by_index]) # sort in ascending order
                
                else:
                    data = sorted(data, key=lambda x: -x[sort_by_index]) # sort in descending order

                # Create an excel sheet
                workbook = xl.Workbook()
                worksheet = workbook.active
                worksheet.title = str(today)
                worksheet.append(["Ім'я", "Виконані цілі", "Заохочення",  "Не виконані цілі", "Всього цілей"])
                # Make the headers bold
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)

                # Add data to the excel sheet
                for i in data:
                    worksheet.append(list(i))

                workbook.save(filename)

                # Add final summary data
                for i in final_weekly_data:
                    worksheet.append(["Підсумок"] + list(i))

                # Make the summary row bold
                last_row = len(worksheet["A"])

                for cell in worksheet[last_row]:
                    cell.font = Font(bold=True)

                workbook.save(filename)

            else: 
                workbook = xl.Workbook()
                worksheet = workbook.active
                worksheet.append(["Ім'я", "Виконані цілі", "Заохочення",  "Не виконані цілі", "Всього цілей"])
                # Make the headers bold
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)

                workbook.save(filename)
    
        elif mode == "month":
            insert_values = (year, month)
            monthly_data_query = """
                SELECT 
                    year, month, week_num, name, completed_goals, rewards, uncompleted_goals, total_goals
                FROM 
                    goals
                WHERE 
                    year = %s and month = %s
                ORDER BY
                    {} {}
            """.format(sort_by, sort_by_direction)

            total_monthly_query = """
                SELECT 
                    year, month, name, sum(completed_goals) as completed_goals , sum(rewards) as rewards, sum(uncompleted_goals) as uncompleted_goals, sum(total_goals) as total_goals
                FROM 
                    goals
                WHERE 
                    year = %s and month = %s
                GROUP BY
                    name, month, year
                ORDER BY
                    {}  {}
            """.format(sort_by, sort_by_direction)

            final_summary_query = """
                SELECT 
                    sum(completed_goals) , sum(rewards), sum(uncompleted_goals), sum(total_goals)
                FROM       
                    goals
                WHERE 
                    year = %s and month = %s
                GROUP BY
                    year, month
            """

            self.__cur.execute(monthly_data_query, insert_values)    
            monthly_data = self.__cur.fetchall()

            self.__cur.execute(total_monthly_query, insert_values)    
            total_data = self.__cur.fetchall()

            self.__cur.execute(final_summary_query, insert_values)
            final_summary_data = self.__cur.fetchall()

            # Create an excel sheet
            workbook = xl.Workbook()

            if monthly_data: 
                worksheet = workbook.active
                year = monthly_data[0][0]
                month = self.__index_to_month[monthly_data[0][1]]
                week_num = monthly_data[0][2]
                sheet_name = f"{year} {month} {week_num}"
                worksheet.title = sheet_name # e.g. 2023 February 3
                worksheet.append(["Ім'я", "Виконані цілі", "Заохочення",  "Не виконані цілі", "Всього цілей"])
                # Make the headers bold
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)

                for i in monthly_data:
                    year = i[0]
                    month = self.__index_to_month[i[1]]
                    week_num = i[2]
                    sheet_name = f"{year} {month} {week_num}"

                    if sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        worksheet.append(list(i)[3:])

                    else:
                        workbook.create_sheet(sheet_name)
                        worksheet = workbook[sheet_name]
                        worksheet.append(["Ім'я", "Виконані цілі", "Заохочення",  "Не виконані цілі", "Всього цілей"])
                        # Make the headers bold
                        for cell in worksheet[1]:
                            cell.font = Font(bold=True)
                        worksheet.append(list(i)[3:])      

                workbook._sheets.sort(key=lambda ws: ws.title)


                # Add total data
                year = total_data[0][0]
                month = self.__index_to_month[total_data[0][1]]
                workbook.create_sheet(f"{year} {month} Підсумки")
                worksheet = workbook[f"{year} {month} Підсумки"]
                worksheet.append(["Ім'я", "Виконані цілі", "Заохочення",  "Не виконані цілі", "Всього цілей"])
                
                # Make the headers bold
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)

                for i in total_data:
                    worksheet.append(list(i[2:]))

                # Add final summary data
                for i in final_summary_data:
                    worksheet.append(["Підсумок"] + list(i))

                # Make the summary row bold
                last_row = len(worksheet["A"])

                for cell in worksheet[last_row]:
                    cell.font = Font(bold=True)

                filename = f"{month} {year}.xlsx"
                workbook.save(filename)

            else: 
                worksheet = workbook.active
                worksheet.append(["Ім'я", "Виконані цілі", "Заохочення",  "Не виконані цілі", "Всього цілей"])
                # Make the headers bold
                for cell in worksheet[1]:
                    cell.font = Font(bold=True)

                filename = f"{month_str} {year}.xlsx"
                workbook.save(filename)

        elif mode == "year":
            pass

        workbook.save(filename)
        return filename


def week_number_of_month(date_value):
     return (date_value.isocalendar()[1] - date_value.replace(day=2).isocalendar()[1] + 1)